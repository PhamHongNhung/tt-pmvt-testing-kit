import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Test Cases"

# ── Màu sắc ──────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F3864"   # xanh đậm
COLOR_HEADER_FONT = "FFFFFF"   # trắng
COLOR_HIGH        = "FF4C4C"   # đỏ nhạt
COLOR_MEDIUM      = "FFB347"   # cam
COLOR_LOW         = "5DBB63"   # xanh lá
COLOR_ROW_ODD     = "F2F7FF"   # xanh dương rất nhạt
COLOR_ROW_EVEN    = "FFFFFF"   # trắng

# ── Dữ liệu ──────────────────────────────────────────────────────────────────
headers = [
    "TC ID", "Module", "Risk Level", "Test Scenario",
    "Pre-Condition", "Test Steps", "Test Data",
    "Expected Result", "Priority"
]

rows = [
    [
        "MYDIO_PAYMENT_PACKAGES_TC_001",
        "Payment > Packages",
        "High",
        "Lấy danh sách gói cước thành công với đầy đủ tham số hợp lệ",
        "Có Bearer Token hợp lệ; server đang hoạt động",
        "1. Gửi GET request tới https://mydio.vn/v1/payment/public/packages?paymentChannel=mps&device_id=NWQ4Yjg3Y2UtODgyNi00YWJh...\n2. Đính kèm header Authorization: Bearer valid_token_abc123\n3. Ghi lại HTTP status code và body response",
        "paymentChannel=mps\ndevice_id=NWQ4Yjg3Y2Ut...\nAuthorization: Bearer valid_token_abc123",
        "HTTP 200 OK; Body là JSON array chứa ít nhất 1 gói cước; Mỗi item có đủ các field: id, name, price, duration",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_002",
        "Payment > Packages",
        "High",
        "Trả về 401 khi không có Authorization header",
        "Server đang hoạt động",
        "1. Gửi GET request với paymentChannel=mps&device_id=NWQ4Yjg3Y2Ut...\n2. Không đính kèm header Authorization\n3. Ghi lại HTTP status code và body",
        "paymentChannel=mps\nKhông có header Authorization",
        "HTTP 401 Unauthorized; Body chứa thông báo lỗi xác thực (VD: {\"error\": \"Unauthorized\"})",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_003",
        "Payment > Packages",
        "High",
        "Trả về 401 khi Bearer Token hết hạn (expired)",
        "Có expired token đã hết hạn; server đang hoạt động",
        "1. Gửi GET request tới endpoint với đầy đủ params hợp lệ\n2. Đính kèm header Authorization: Bearer expired_token_xyz\n3. Ghi lại HTTP status code và body",
        "Authorization: Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.EXPIRED.signature\npaymentChannel=mps",
        "HTTP 401 Unauthorized; Body chứa message lỗi liên quan đến token hết hạn",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_004",
        "Payment > Packages",
        "High",
        "Trả về 401 khi Bearer Token sai định dạng / giả mạo",
        "Server đang hoạt động",
        "1. Gửi GET request với Authorization: Bearer INVALID_FAKE_TOKEN_000\n2. Ghi lại HTTP status code và body",
        "Authorization: Bearer INVALID_FAKE_TOKEN_000\npaymentChannel=mps",
        "HTTP 401 Unauthorized; Body trả về thông báo lỗi; Không trả về dữ liệu gói cước",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_005",
        "Payment > Packages",
        "High",
        "Trả về lỗi hoặc danh sách rỗng khi paymentChannel không hợp lệ",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request với paymentChannel=INVALID_CHANNEL_XYZ\n2. Đính kèm header Authorization hợp lệ\n3. Ghi lại HTTP status code và body",
        "paymentChannel=INVALID_CHANNEL_XYZ\ndevice_id=NWQ4Yjg3Y2Ut...\nAuthorization: Bearer valid_token_abc123",
        "HTTP 400 Bad Request HOẶC HTTP 200 với body là mảng rỗng []; Không trả về dữ liệu gói cước",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_006",
        "Payment > Packages",
        "Medium",
        "Trả về lỗi khi bỏ qua tham số paymentChannel (thiếu param bắt buộc)",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request tới https://mydio.vn/v1/payment/public/packages (không có query param)\n2. Đính kèm Authorization hợp lệ\n3. Ghi lại response",
        "Không có paymentChannel\nKhông có device_id\nAuthorization: Bearer valid_token_abc123",
        "HTTP 400 Bad Request; Body chứa thông báo yêu cầu thiếu tham số; HOẶC HTTP 200 với danh sách fallback",
        "Medium"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_007",
        "Payment > Packages",
        "Medium",
        "Lấy danh sách gói cước thành công khi không có device_id (optional param)",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request tới .../packages?paymentChannel=mps\n2. Không truyền device_id\n3. Đính kèm Authorization hợp lệ\n4. Ghi lại response",
        "paymentChannel=mps\ndevice_id không có\nAuthorization: Bearer valid_token_abc123",
        "HTTP 200 OK; Body là JSON array hợp lệ; Hệ thống không báo lỗi vì thiếu device_id",
        "Medium"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_008",
        "Payment > Packages",
        "Medium",
        "Xử lý đúng khi device_id là chuỗi rỗng",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request với paymentChannel=mps&device_id= (truyền rỗng)\n2. Đính kèm Authorization hợp lệ\n3. Ghi lại response",
        "paymentChannel=mps\ndevice_id= (empty string)\nAuthorization: Bearer valid_token_abc123",
        "HTTP 200 OK (trả dữ liệu bình thường) HOẶC HTTP 400 nếu device_id là bắt buộc; Không crash server (không phải 500)",
        "Medium"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_009",
        "Payment > Packages",
        "Medium",
        "Xử lý đúng khi device_id chứa ký tự đặc biệt không encode (XSS payload)",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request với device_id=<script>alert(1)</script> (không encode)\n2. Đính kèm Authorization hợp lệ\n3. Ghi lại response",
        "paymentChannel=mps\ndevice_id=<script>alert(1)</script>\nAuthorization: Bearer valid_token_abc123",
        "HTTP 400 HOẶC HTTP 200 nhưng server sanitize dữ liệu; Tuyệt đối không thực thi script hay trả về lỗi 500",
        "Medium"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_010",
        "Payment > Packages",
        "High",
        "Decision Table: Token hợp lệ + paymentChannel hợp lệ → 200 OK",
        "Có Token hợp lệ; paymentChannel=mps",
        "1. Gửi GET request đầy đủ hợp lệ\n2. Ghi lại response",
        "Authorization: Bearer valid_token_abc123\npaymentChannel=mps",
        "HTTP 200; Có data JSON hợp lệ",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_011",
        "Payment > Packages",
        "High",
        "Decision Table: Token hợp lệ + paymentChannel không hợp lệ → lỗi",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request với paymentChannel=FAKECHANNEL99\n2. Đính kèm Authorization hợp lệ",
        "paymentChannel=FAKECHANNEL99\nAuthorization: Bearer valid_token_abc123",
        "HTTP 400 HOẶC 200 với mảng rỗng; Không có dữ liệu gói cước",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_012",
        "Payment > Packages",
        "High",
        "Decision Table: Token không hợp lệ + paymentChannel hợp lệ → 401",
        "Token không hợp lệ",
        "1. Gửi GET request với paymentChannel=mps\n2. Đính kèm Authorization: Bearer WRONG_TOKEN_999",
        "paymentChannel=mps\nAuthorization: Bearer WRONG_TOKEN_999",
        "HTTP 401 Unauthorized; Không trả về dữ liệu",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_013",
        "Payment > Packages",
        "Low",
        "Kiểm tra header cache-control: no-store được tuân thủ trong response",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request hợp lệ\n2. Kiểm tra response headers\n3. Xác nhận không có Cache-Control: public hay Expires trong response",
        "paymentChannel=mps\nAuthorization: Bearer valid_token_abc123\ncache-control: no-store trong request",
        "Response header không chứa Cache-Control: public; Dữ liệu gói cước không bị cache",
        "Low"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_014",
        "Payment > Packages",
        "High",
        "Kiểm tra cấu trúc JSON response đúng schema (Schema Validation)",
        "Có Bearer Token hợp lệ",
        "1. Gửi GET request hợp lệ\n2. Parse JSON response\n3. Validate từng gói cước có đủ fields: id, name, price, currency, duration",
        "paymentChannel=mps\nAuthorization: Bearer valid_token_abc123",
        "HTTP 200; Mỗi object có đủ fields bắt buộc; price là number dương; currency là string hợp lệ (VD: VND); không có field null nếu là bắt buộc",
        "High"
    ],
    [
        "MYDIO_PAYMENT_PACKAGES_TC_015",
        "Payment > Packages",
        "Low",
        "Kiểm tra API phản hồi trong thời gian chấp nhận được (Performance)",
        "Server hoạt động bình thường; mạng ổn định",
        "1. Gửi GET request hợp lệ\n2. Đo thời gian từ lúc gửi đến lúc nhận response hoàn toàn",
        "paymentChannel=mps\nAuthorization: Bearer valid_token_abc123",
        "Thời gian phản hồi ≤ 2000ms (2 giây); HTTP 200; Body hợp lệ",
        "Low"
    ],
]

# ── Borders ───────────────────────────────────────────────────────────────────
thin = Side(style='thin', color="CCCCCC")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Header row ────────────────────────────────────────────────────────────────
header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_font = Font(bold=True, color=COLOR_HEADER_FONT, name="Calibri", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, col_name in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = cell_border

ws.row_dimensions[1].height = 30

# ── Risk color map ────────────────────────────────────────────────────────────
risk_colors = {
    "High":   ("FFE5E5", "C00000"),
    "Medium": ("FFF3E0", "E65100"),
    "Low":    ("E8F5E9", "1B5E20"),
}

prio_colors = {
    "High":   ("C00000", "FFFFFF"),
    "Medium": ("E65100", "FFFFFF"),
    "Low":    ("1B5E20", "FFFFFF"),
}

# ── Data rows ────────────────────────────────────────────────────────────────
cell_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, row_data in enumerate(rows, start=2):
    risk_val = row_data[2]
    prio_val = row_data[8]
    bg_color = risk_colors.get(risk_val, ("FFFFFF", "000000"))[0]
    row_fill = PatternFill("solid", fgColor=COLOR_ROW_ODD if row_idx % 2 == 0 else COLOR_ROW_EVEN)

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10)
        cell.border = cell_border

        # Risk Level cell styling
        if col_idx == 3:
            rc = risk_colors.get(value, ("FFFFFF", "000000"))
            cell.fill = PatternFill("solid", fgColor=rc[0])
            cell.font = Font(name="Calibri", size=10, bold=True, color=rc[1])
            cell.alignment = center_align
        # Priority cell styling
        elif col_idx == 9:
            pc = prio_colors.get(value, ("FFFFFF", "000000"))
            cell.fill = PatternFill("solid", fgColor=pc[0])
            cell.font = Font(name="Calibri", size=10, bold=True, color=pc[1])
            cell.alignment = center_align
        # TC ID – bold
        elif col_idx == 1:
            cell.fill = PatternFill("solid", fgColor="EBF3FB")
            cell.font = Font(name="Calibri", size=10, bold=True, color="1F3864")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        else:
            cell.fill = row_fill
            cell.alignment = cell_align

    ws.row_dimensions[row_idx].height = 90

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [35, 20, 12, 45, 35, 60, 45, 55, 10]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Freeze header ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Auto filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "/Users/hongnhung/Desktop/Testing/Nhung_Auto/.agent/apis/testcases_payment_packages.xlsx"
wb.save(out_path)
print(f"✅ Đã tạo file Excel: {out_path}")
